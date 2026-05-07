from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List


MAX_SHEETS = 50
MAX_ROWS_PER_SHEET = 5000
MAX_COLS_PER_SHEET = 200


def build_xlsx_output(
    full_text: str,
    sheets: List[Dict[str, Any]],
    meta: Dict[str, Any],
    warnings: List[Dict[str, str]],
) -> Dict[str, Any]:
    return {
        "full_text": full_text,
        "pages": sheets,
        "meta": meta,
        "warnings": warnings,
    }


def empty_xlsx_output(meta_overrides: Dict[str, Any], warnings: List[Dict[str, str]]) -> Dict[str, Any]:
    base_meta = {
        "source": "xlsx",
        "engine": "openpyxl",
        "page_count": 0,
        "title": None,
        "author": None,
        "created_at": None,
        "is_encrypted": False,
        "avg_confidence": 1.0,
    }
    base_meta.update(meta_overrides)
    return build_xlsx_output("", [], base_meta, warnings)


def format_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value).strip()


def extract_sheet(worksheet: Any, sheet_index: int) -> Dict[str, Any]:
    rows_text: List[str] = []
    row_count = 0
    truncated_rows = False

    for row in worksheet.iter_rows(values_only=True):
        if row_count >= MAX_ROWS_PER_SHEET:
            truncated_rows = True
            break
        cells = [format_cell_value(cell) for cell in row[:MAX_COLS_PER_SHEET]]
        if any(cells):
            rows_text.append("\t".join(cells).rstrip("\t"))
        row_count += 1

    sheet_text = "\n".join(rows_text).strip()
    label = f"[Sheet: {worksheet.title}]"
    decorated = f"{label}\n{sheet_text}" if sheet_text else label

    return {
        "page_number": sheet_index,
        "sheet_name": worksheet.title,
        "text": decorated,
        "row_count": row_count,
        "truncated": truncated_rows,
        "blocks": [],
    }


def extract_xlsx_text(file_path: Path) -> Dict[str, Any]:
    warnings: List[Dict[str, str]] = []

    if not file_path.exists():
        warnings.append({"code": "XLSX-MISSING", "message": "File not found at the expected path."})
        return empty_xlsx_output({}, warnings)

    if file_path.stat().st_size == 0:
        warnings.append({"code": "XLSX-EMPTY-FILE", "message": "File is zero bytes."})
        return empty_xlsx_output({}, warnings)

    try:
        from openpyxl import load_workbook
    except ImportError:
        warnings.append(
            {"code": "XLSX-DEP-MISSING", "message": "openpyxl is not installed; XLSX parsing is unavailable."}
        )
        return empty_xlsx_output({}, warnings)

    try:
        workbook = load_workbook(filename=str(file_path), read_only=True, data_only=True)
    except Exception as exc:
        message = str(exc).lower()
        if "encrypted" in message or "password" in message:
            warnings.append(
                {"code": "XLSX-ENCRYPTED", "message": "Workbook is password-protected and could not be opened."}
            )
            return empty_xlsx_output({"is_encrypted": True}, warnings)

        warnings.append(
            {"code": "XLSX-CORRUPTED", "message": f"openpyxl could not open the workbook: {exc.__class__.__name__}."}
        )
        return empty_xlsx_output({}, warnings)

    try:
        sheet_names = workbook.sheetnames
        sheet_count = len(sheet_names)
        if sheet_count > MAX_SHEETS:
            warnings.append(
                {
                    "code": "XLSX-SHEET-LIMIT",
                    "message": f"Workbook has {sheet_count} sheets; only the first {MAX_SHEETS} were processed.",
                }
            )
            sheet_names = sheet_names[:MAX_SHEETS]

        sheets: List[Dict[str, Any]] = []
        full_text_parts: List[str] = []

        for sheet_index, sheet_name in enumerate(sheet_names, start=1):
            try:
                worksheet = workbook[sheet_name]
                sheet_data = extract_sheet(worksheet, sheet_index)
            except Exception as exc:
                warnings.append(
                    {
                        "code": "XLSX-SHEET-ERROR",
                        "message": f"Sheet '{sheet_name}' could not be read ({exc.__class__.__name__}); skipped.",
                    }
                )
                continue

            sheets.append(sheet_data)
            full_text_parts.append(sheet_data["text"])
            if sheet_data["truncated"]:
                warnings.append(
                    {
                        "code": "XLSX-ROW-LIMIT",
                        "message": (
                            f"Sheet '{sheet_name}' has more than {MAX_ROWS_PER_SHEET} rows; "
                            "later rows were not extracted."
                        ),
                    }
                )

        full_text = "\n\n".join(part for part in full_text_parts if part).strip()

        if not full_text:
            warnings.append(
                {"code": "XLSX-NO-TEXT", "message": "Workbook contains no readable cell content."}
            )

        properties = workbook.properties
        meta = {
            "source": "xlsx",
            "engine": "openpyxl",
            "avg_confidence": 1.0,
            "page_count": len(sheets),
            "title": (getattr(properties, "title", None) or "").strip() or None,
            "author": (getattr(properties, "creator", None) or "").strip() or None,
            "created_at": getattr(properties, "created", None).isoformat() if getattr(properties, "created", None) else None,
            "is_encrypted": False,
        }

        return build_xlsx_output(full_text, sheets, meta, warnings)
    finally:
        workbook.close()


if __name__ == "__main__":
    import json
    import sys

    if len(sys.argv) < 2:
        print("Usage: python xlsx_extractor.py <path-to-xlsx>")
        raise SystemExit(1)

    result = extract_xlsx_text(Path(sys.argv[1]))
    print(json.dumps(result, indent=2, ensure_ascii=False, default=str))
